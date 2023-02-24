"""
OpenNEM NEM General Information Parser

This module parses the general information Excel files that are released
by AEMO (almost) every month. The downloads can be found on this page:


https://www.aemo.com.au/energy-systems/electricity/national-electricity-market-nem/nem-forecasting-and-planning/forecasting-and-planning-data/generation-information

In opennem.spiders.aemo.monitoring there is a spider that finds the new files
and they will be passed into this file to be output into a StationSet schema,
ready for manipulation, export into different formats - or, for our primary
purpose - to import new facilities and updates into the OpenNEM database.
"""

import logging
import re
from datetime import date
from pathlib import Path
from tempfile import mkdtemp
from typing import Any

from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from pydantic import validator

from opennem.core.normalizers import clean_float, normalize_duid, station_name_cleaner
from opennem.schema.core import BaseConfig
from opennem.settings import settings
from opennem.utils.dates import get_today_opennem
from opennem.utils.http import download_file

logger = logging.getLogger("opennem.parsers.aemo.gi")


AEMO_GI_DOWNLOAD_URL = (
    "https://www.aemo.com.au/-/media/files/electricity/nem/planning_and_forecasting/"
    "generation_information/2022/nem-generation-information-{month}-{year}.xlsx?la=en"
)


class OpennemAEMOGIParserException(Exception):
    """Exception raised in this module"""

    pass


GI_EXISTING_NEW_GEN_KEYS = {
    "region": "A",
    "AssetType": "B",
    "StationName": "C",
    "Owner": "D",
    "TechType": "E",
    "FuelType": "F",
    "duid": "G",
    "units_no": "H",
    "capacity_registered": "M",
    "StorageCapacity": "N",
    "UnitStatus": "O",
    "DispatchType": "P",
    "UseDate": "Q",
    "ClosureDateExpected": "R",
    "ClosureDate": "S",
    "SurveyID": "V",
    "FuelSummary": "U",
    "SurveyEffective": "Y",
}


def excel_column_to_column_index(excel_column: str) -> int:
    """Takes an excel column like 'A' and converts to column index"""
    return ord(excel_column.upper()) - 64


AEMO_GI_FUELTECH_MAP = {
    "Solar": "solar_utility",
    "Battery Storage": "battery_charging",
    "Coal": "coal_black",
    "CCGT": "gas_ccgt",
    "Water": "hydo",
    "Wind": "wind",
    "Biomass": "bioenergy_biomass",
    "OCGT": "gas_ocgt",
}

AEMO_GI_STATUS_MAP = {
    "Anticipated": None,
    "Committed": "committed",
    "Publicly Announced Upgrade": None,
    "Committed¹": "committed",
    "Committed Upgrade": None,
    "Withdrawn - Permanent": None,
    "Committed*": "committed",
    "In Service - Announced Withdrawal (Permanent)": "operating",
    "In Commissioning": "commissioning",
    "In Service": "operating",
    "Publicly Announced": "announced",
    "Anticipated Upgrade": None,
}


def aemo_gi_fueltech_to_fueltech(gi_fueltech: str | None) -> str | None:
    """Map AEMO GI fueltech to OpenNEM fueltech"""
    if not gi_fueltech:
        return None

    if gi_fueltech not in AEMO_GI_FUELTECH_MAP.keys():
        return None

    return AEMO_GI_FUELTECH_MAP[gi_fueltech]


def aemo_gi_status_map(gi_status: str | None) -> str | None:
    """Map AEMO GI status to OpenNEM status"""
    if not gi_status:
        return None

    if gi_status not in AEMO_GI_STATUS_MAP.keys():
        return None

    return AEMO_GI_STATUS_MAP[gi_status]


def aemo_gi_capacity_cleaner(cap: str | None) -> float | None:
    """Custom capacity cleaner because sometimes its parsed as silly
    text like a range (ie. '150 - 180'"""
    if isinstance(cap, int) or isinstance(cap, float):
        return cap

    if not cap:
        return None

    cap = cap.strip()

    num_part = re.search(r"^[\d\.]+", cap)

    if not num_part:
        return None

    num_extracted = str(num_part.group(0))

    num_extracted_and_clean = clean_float(num_extracted)

    return num_extracted_and_clean


class AEMOGIRecord(BaseConfig):
    name: str
    name_network: str
    region: str
    fueltech_id: str | None
    status_id: str | None
    duid: str | None
    units_no: int | None
    capacity_registered: float | None

    _clean_duid = validator("duid", pre=True, allow_reuse=True)(normalize_duid)
    _clean_capacity = validator("capacity_registered", pre=True, allow_reuse=True)(aemo_gi_capacity_cleaner)


def parse_aemo_general_information(filename: Path) -> list[AEMOGIRecord]:
    """Primary record parser for GI information. Takes the spreadsheet location from a path,
    parses it and the relevant sheet and returns a list of GI records"""

    wb = load_workbook(str(filename), data_only=True)

    SHEET_KEY = "ExistingGeneration&NewDevs"

    if SHEET_KEY not in wb:
        raise OpennemAEMOGIParserException("Doesn't look like a GI spreadsheet")

    ws = wb[SHEET_KEY]

    records = []

    for row in ws.iter_rows(min_row=3, values_only=True):
        # pick out the columns we want
        # lots of hidden columns in the sheet
        row_collapsed = [row[excel_column_to_column_index(i) - 1] for i in GI_EXISTING_NEW_GEN_KEYS.values()]

        return_dict = dict(zip(GI_EXISTING_NEW_GEN_KEYS, list(row_collapsed)))

        # break at end of data records
        # GI has a blank line before garbage notes
        if row[0] is None:
            break

        if return_dict is None:
            raise OpennemAEMOGIParserException(f"Failed on row: {row}")

        if return_dict["region"] is None or return_dict["region"] not in ["NSW1", "QLD1", "SA1", "TAS1", "VIC1"]:
            continue

        return_dict = {
            **return_dict,
            **{
                "name": station_name_cleaner(return_dict["StationName"]),
                "name_network": return_dict["StationName"],
                "status_id": aemo_gi_status_map(return_dict["UnitStatus"]),
                "fueltech_id": aemo_gi_fueltech_to_fueltech(return_dict["FuelSummary"]),
            },
        }

        return_model = AEMOGIRecord(**return_dict)

        records.append(return_model)

    logger.info(f"Parsed {len(records)} records")

    return records


def get_unique_values_for_field(records: list[dict], field_name: str) -> list[Any]:
    return list({i[field_name] for i in records})


def get_aemo_gi_download_url(month: date) -> str:
    """Get the AEMO GI download URL for a given month"""
    return AEMO_GI_DOWNLOAD_URL.format(month=month.strftime("%B").lower(), year=month.strftime("%Y"))


def download_latest_aemo_gi_file() -> Path:
    """This will download the latest GI file into a local temp directory
    and return the path to it"""

    gi_saved_path: Path | None = None
    dest_dir = Path(mkdtemp(prefix=f"{settings.tmp_file_prefix}"))

    this_month = get_today_opennem()
    last_month = this_month - relativedelta(months=1)

    for download_date in [this_month, last_month]:
        download_url = get_aemo_gi_download_url(download_date)

        logger.info(f"Downloading AEMO GI file from {download_url}")

        try:
            gi_saved_path = download_file(
                download_url, dest_dir, expect_content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            if gi_saved_path:
                return gi_saved_path
        except Exception as e:
            logger.error(f"Failed to download AEMO GI file: {e}")
            continue

    if not gi_saved_path:
        raise OpennemAEMOGIParserException("Failed to download AEMO GI file")

    return gi_saved_path


def download_and_parse_aemo_gi_file() -> list[AEMOGIRecord]:
    """This will download the latest GI file and parse it"""
    gi_latest_path = download_latest_aemo_gi_file()
    records = parse_aemo_general_information(gi_latest_path)

    return records


# debug entrypoint
if __name__ == "__main__":
    records = download_and_parse_aemo_gi_file()

    for r in records:
        print(f"{r.name_network} -> {r.capacity_registered}")
